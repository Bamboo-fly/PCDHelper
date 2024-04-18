import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取第一个表格
df1 = pd.read_excel(r"C:\Users\Bamboo\Desktop\科学院.xlsx", engine='openpyxl', header=None)

# 读取第二个表格
df2 = pd.read_excel(r"C:\Users\Bamboo\Desktop\国家奖初评专家名单.xlsx", engine='openpyxl', header=None)

# 创建一个黄色的填充样式
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 获取第二个表格的 Excel 文件对象
wb = load_workbook(r"C:\Users\Bamboo\Desktop\国家奖初评专家名单.xlsx")
ws = wb.active

# 获取第一个表格的所有人名
names_df1 = set(df1[0])
num=0
# 在第二个表格中查找第一个表格中存在的人名
for index, row in df2.iterrows():
    name = row[0]
    if name in names_df1:
        num+=1
        print(num)
        # 获取单元格的坐标
        cell = ws.cell(row=index + 1, column=1)
        # 标记为黄色
        cell.fill = yellow_fill

# 另存为标记后的 Excel 文件
wb.save(r"C:\Users\Bamboo\Desktop\标记后科学院名单.xlsx")
